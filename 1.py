import fitz
from PIL import Image
import cv2
import re
import numpy as np
from paddleocr import PPStructure, PaddleOCR
import os
import openpyxl
from openpyxl.utils import get_column_letter

import pandas as pd
from pathlib import Path


def process_pdf(pdf_path):
    pattern = r'^[a-zA-Z,\s\-\']+$'

    def is_valid(text):
        return bool(re.match(pattern, text))

    def preprocess_image(image):
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1]
        return thresh

    # 初始化PPStructure和PaddleOCR模型
    table_engine = PPStructure(show_log=True)
    ocr = PaddleOCR(use_angle_cls=True, lang="en", use_gpu=False)

    # 打开PDF文件
    pdf_document = fitz.open(pdf_path)

    # 获取第一页
    page = pdf_document[0]

    # 将页面渲染为图像
    pix = page.get_pixmap(matrix=fitz.Matrix(4, 3))
    img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
    img_np = np.array(img)

    # 预处理图像
    processed_img = preprocess_image(img_np)

    # 保存为临时图像文件
    temp_image_path = "temp_image.png"
    cv2.imwrite(temp_image_path, processed_img)

    # 使用PPStructure进行表格识别
    result = table_engine(temp_image_path)

    data = {'date': '', 'list': []}
    re_ss = []
    s_index = 1

    ocr_result = ocr.ocr(temp_image_path, cls=True)

    for idx, line in enumerate(ocr_result):
        i_index = 0
        for word_info in line:
            print(word_info[1][0])

            i_index += 1
            if word_info[1][0] == "1REMARK":
                break
            if i_index == 7:
                data['date'] = word_info[1][0]
                continue
            if i_index >= 17:
                if is_valid(word_info[1][0]):
                    new_title = word_info[1][0]
                    new_title = new_title.replace(",", ",")
                    if i_index > 17:
                        if ',' in new_title:
                            print("###########xxxx################")
                            print(new_title)
                            print("############uuuu###############")
                            data['list'].append(re_ss)
                            re_ss = []
                    re_ss.append(new_title.strip())
                else:
                    re_title = word_info[1][0].strip()
                    if ":" in re_title:
                        parts = re_title.split(':')
                        re_title = f"{parts[0].zfill(2)}:{parts[1].zfill(2)}"
                    re_ss.append(re_title)

    # 添加最后一个 re_ss 到 data['list']
    if re_ss:
        data['list'].append(re_ss)

    # 清理临时文件
    os.remove(temp_image_path)

    # 关闭PDF文档
    pdf_document.close()

    return data


def add_row_and_insert_data(file_path, new_data):
    # 加载工作簿
    wb = openpyxl.load_workbook(file_path)

    # 选择活动工作表
    ws = wb.active

    # 找到第二列（B列）的最后一个非空单元格
    last_row = ws.max_row
    while last_row > 0 and ws.cell(row=last_row, column=2).value is None:
        last_row -= 1

    # 在最后一个非空单元格后新增一行
    new_row = last_row + 1

    # 从新行的第二列（B列）开始插入数据
    for i, value in enumerate(new_data):
        column = get_column_letter(2 + i)  # 从B列开始
        ws[f"{column}{new_row}"] = value

    # 保存更改
    wb.save(file_path)
    print(f"数据已添加到第 {new_row} 行，从B列开始")


# 使用示例


def search_excel_file(file_path, search_strings, columns_to_search):
    """
    在Excel文件中搜索指定列的匹配行。

    参数:
    file_path (str or Path): Excel文件的路径
    search_strings (list): 要搜索的字符串列表，顺序对应columns_to_search
    columns_to_search (list): 要搜索的列名列表

    返回:
    DataFrame: 包含匹配行的DataFrame，如果没有匹配则为空DataFrame
    """
    try:
        # 将文件路径转换为Path对象
        file_path = Path(file_path)

        # 读取 Excel 文件
        df = pd.read_excel(file_path)

        # 确保指定的列存在
        if not all(col in df.columns for col in columns_to_search):
            print("指定的列不存在于 Excel 文件中")
            return pd.DataFrame()

        # 清理数据：将指定列转换为字符串并去除前后空格
        for col in columns_to_search:
            df[col] = df[col].astype(str).str.strip()

        # 构建查询条件
        query_condition = True
        for col, search_string in zip(columns_to_search, search_strings):
            query_condition &= (df[col] == search_string)

        # 进行查询
        result = df[query_condition]

        return result

    except Exception as e:
        print(f"发生错误: {e}")
        return pd.DataFrame()


# 使用示例
if __name__ == "__main__":

    # file_path = Path(r'./doc/8-19-2024 Burchett C-10003 Moss Mountain.xlsx')
    # print(file_path)
    #
    # new_data = ['Data 1', 'Data 2', 'Data 3', 'Data 4']  # 要插入的四列数据
    #
    # add_row_and_insert_data(file_path, new_data)

    # 使用示例
    result = process_pdf("./doc/20240820200038923.pdf")
    columns_to_search = ['First Name', 'Middle Name', 'Last Name']
    file_path = './doc/Change Out Report 2024.xlsx'
    file_path_2 = Path(r'./doc/8-19-2024 Burchett C-10003 Moss Mountain.xlsx')

    print(result)

    for info in result['list']:
        # print(info)
        if info[0]:
            print(info[0])
            original_string = info[0]
            # 使用 split() 方法拆分字符串，然后使用列表推导去除空白字符
            name_parts = [part.strip() for part in original_string.split(',')]
            # 将所有部分连接成一个字符串，然后再次分割
            full_name = ' '.join(name_parts)
            search_strings = full_name.split()
            print(search_strings)
            # search_strings = ['Austin', 'Thomas', 'Hartelt']

            result = search_excel_file(file_path, search_strings, columns_to_search)
            if result.empty:
                print("没有找到匹配的行")
            else:
                print("找到以下匹配的行：")
                new_data = info  # 要插入的四列数据
                del new_data[1]
                add_row_and_insert_data(file_path_2, new_data)
                print(result)

import fitz
from PIL import Image
import cv2
import re
import numpy as np
from paddleocr import PPStructure, PaddleOCR
import os
import openpyxl
from openpyxl.utils import get_column_letter
import pandas as pd
from pathlib import Path


class PDFProcessor:
    def __init__(self, pdf_path, excel_file_path, output_file_path):
        self.pdf_path = pdf_path
        self.excel_file_path = excel_file_path
        self.output_file_path = Path(output_file_path)
        self.table_engine = PPStructure(show_log=True)
        self.ocr = PaddleOCR(use_angle_cls=True, lang="en", use_gpu=False)
        self.columns_to_search = ['First Name', 'Middle Name', 'Last Name']

    @staticmethod
    def preprocess_image(image):
        gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1]
        return thresh

    @staticmethod
    def is_valid_name(text):
        pattern = r'^[a-zA-Z,\s\-\.\!\']+$'
        return bool(re.match(pattern, text))

    @staticmethod
    def remove_dots(text):
        return re.sub(r'^\.|\.+$', '', text)

    @staticmethod
    def remove_exclamations(s):
        return s.strip('!')

    @staticmethod
    def format_time(time_str):
        if ":" in time_str:
            parts = time_str.split(':')
            return f"{parts[0].zfill(2)}:{parts[1].zfill(2)}"
        if time_str.isdigit():
            # 如果长度小于4，在前面补0
            s = time_str.zfill(4)

            # 只取前四位
            s = s[:4]

            # 用冒号分割
            return f"{s[:2]}:{s[2:]}"
            # 在中间插入冒号
            # return f"{time_str[:2]}:{time_str[2:]}"

        return time_str

    def process_pdf(self):
        with fitz.open(self.pdf_path) as pdf_document:
            page = pdf_document[0]
            pix = page.get_pixmap(matrix=fitz.Matrix(3, 3))
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
            img_np = np.array(img)

        processed_img = self.preprocess_image(img_np)
        temp_image_path = "temp_image.png"
        cv2.imwrite(temp_image_path, processed_img)

        # result = self.table_engine(temp_image_path)
        data = {'date': '', 'list': []}
        re_ss = []
        kai_start = False
        ocr_result = self.ocr.ocr(temp_image_path, cls=True)
        for line in ocr_result:
            for i, word_info in enumerate(line, start=1):
                text = word_info[1][0]
                text = self.remove_dots(text)
                text = self.remove_exclamations(text)
                text = text.replace(",", ",").strip()
                text = text.replace(" i", "").strip()

                if "," in text:
                    kai_start = True

                print(text)
                if "REMARK" in text:
                    break
                if text == "OFF" or text == "OFE" or text == "ON":
                    continue
                if i == 7:
                    data['date'] = text
                elif i > 17 or kai_start is True:
                    if self.is_valid_name(text):
                        if ',' in text:
                            if len(re_ss) > 0:
                                data['list'].append(re_ss)
                            re_ss = []
                        re_ss.append(text)
                    else:
                        if '|' in text:
                            new_li = text.split('|')
                            for word in new_li:
                                re_ss.append(self.format_time(word.strip()))

                        else:
                            re_ss.append(self.format_time(text.strip()))

        if re_ss:
            data['list'].append(re_ss)

        os.remove(temp_image_path)
        return data

    def add_row_and_insert_data(self, new_data):
        try:
            wb = openpyxl.load_workbook(self.output_file_path)
            ws = wb.active

            last_row = max((c.row for c in ws['B'] if c.value is not None), default=0)
            new_row = last_row + 1

            for i, value in enumerate(new_data, start=2):
                ws[f"{get_column_letter(i)}{new_row}"] = value

            wb.save(self.output_file_path)
            print(f"数据已添加到第 {new_row} 行，从B列开始")

        except Exception as e:
            print(f"添加数据时发生错误: {e}")
        finally:
            if 'wb' in locals():
                wb.close()

    def search_excel_file(self, search_strings):
        try:
            df = pd.read_excel(self.excel_file_path)
            if not all(col in df.columns for col in self.columns_to_search):
                print("指定的列不存在于 Excel 文件中")
                return pd.DataFrame()

            for col in self.columns_to_search:
                df[col] = df[col].astype(str).str.strip()

            query_condition = pd.Series([True] * len(df))
            for col, search_string in zip(self.columns_to_search, search_strings):
                query_condition &= (df[col] == search_string)

            return df[query_condition]
        except Exception as e:
            print(f"搜索Excel文件时发生错误: {e}")
            return pd.DataFrame()

    def process(self):
        result = self.process_pdf()
        print(result)

        for info in result['list']:
            if len(info) > 0 and len(info[0]) > 0:
                name_parts = [part.strip() for part in info[0].split(',')]
                full_name = ' '.join(name_parts)
                search_strings = full_name.split()
                print(search_strings)

                # search_result = self.search_excel_file(search_strings)
                new_data = info
                del new_data[1]
                print(new_data)
                self.add_row_and_insert_data(new_data)
                # if search_result.empty:
                #     print("没有找到匹配的行")
                # else:
                #     print("找到以下匹配的行：")
                #     print(search_result)
                #     new_data = info[:]
                #     del new_data[1]
                #     self.add_row_and_insert_data(new_data)


def main():
    # pdf_path = "./doc/20240820200058571.pdf"
    # pdf_path = "./doc/20240820200116072.pdf"
    pdf_path = "./doc/20240820200038923.pdf"
    excel_file_path = './doc/Change Out Report 2024.xlsx'
    output_file_path = './doc/8-19-2024 Burchett C-10003 Moss Mountain.xlsx'

    processor = PDFProcessor(pdf_path, excel_file_path, output_file_path)
    processor.process()


if __name__ == "__main__":
    main()
